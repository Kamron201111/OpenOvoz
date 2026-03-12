import logging
import os
import re
import io
from datetime import datetime

import requests
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    MessageHandler,
    ConversationHandler,
    filters,
    ContextTypes,
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────── CONFIG ───────────────────
BOT_TOKEN = os.getenv("BOT_TOKEN", "YOUR_BOT_TOKEN_HERE")

# OpenBudget API endpoints
API_BASE = "https://openbudget.uz/api/v1"
VOTES_ENDPOINT = f"{API_BASE}/votes"           # GET  ?initiative_id=...
SCRAPE_ENDPOINT = f"{API_BASE}/scrape_votes"   # POST {"initiative_id": "..."}

HEADERS = {
    "Content-Type": "application/json",
    "Accept": "application/json",
}

# ─────────────────── STATES ───────────────────
WAITING_INITIATIVE_ID = 1
WAITING_COUNT = 2
WAITING_SEARCH_PHONE = 3
WAITING_SEARCH_INITIATIVE = 4

# ─────────────────── LOGGING ──────────────────
logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger(__name__)


# ════════════════════════════════════════════
#  HELPER: Fetch votes from API
# ════════════════════════════════════════════
def fetch_votes(initiative_id: str) -> list[dict] | None:
    """
    Returns list of vote dicts: [{"phoneNumber": "**-*63-33-35", "voteDate": "..."}, ...]
    Returns None on error.
    """
    try:
        # Try POST scrape_votes first
        resp = requests.post(
            SCRAPE_ENDPOINT,
            json={"initiative_id": initiative_id},
            headers=HEADERS,
            timeout=30,
        )
        if resp.status_code == 200:
            data = resp.json()
            if data.get("success") and "votes" in data.get("data", {}):
                return data["data"]["votes"]

        # Fallback: GET votes
        resp = requests.get(
            VOTES_ENDPOINT,
            params={"initiative_id": initiative_id},
            headers=HEADERS,
            timeout=30,
        )
        if resp.status_code == 200:
            data = resp.json()
            if isinstance(data, list):
                return data
            if data.get("success") and "votes" in data.get("data", {}):
                return data["data"]["votes"]
            if "votes" in data:
                return data["votes"]

        logger.warning("API error: %s %s", resp.status_code, resp.text[:200])
        return None

    except Exception as e:
        logger.error("fetch_votes error: %s", e)
        return None


# ════════════════════════════════════════════
#  HELPER: Build Excel file
# ════════════════════════════════════════════
def build_excel(votes: list[dict], initiative_id: str, count: int) -> io.BytesIO:
    """Creates a formatted Excel file and returns it as BytesIO."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ovozlar"

    # ── Styles ──
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F5C99")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    cell_font = Font(name="Arial", size=10)
    cell_align_center = Alignment(horizontal="center", vertical="center")
    cell_align_left = Alignment(horizontal="left", vertical="center")

    alt_fill = PatternFill("solid", start_color="EBF3FB")
    white_fill = PatternFill("solid", start_color="FFFFFF")

    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ──
    ws.merge_cells("A1:C1")
    title_cell = ws["A1"]
    title_cell.value = f"OpenBudget.uz — Ovozlar ro'yxati"
    title_cell.font = Font(name="Arial", bold=True, size=13, color="1F5C99")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:C2")
    sub = ws["A2"]
    sub.value = (
        f"Tashabbus ID: {initiative_id}   |   "
        f"Jami ko'rsatilgan: {count} ta   |   "
        f"Yuklab olingan vaqt: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    )
    sub.font = Font(name="Arial", size=9, italic=True, color="595959")
    sub.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Header row ──
    headers = ["T/R", "Telefon raqam", "Ovoz sanasi"]
    col_widths = [7, 22, 24]

    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.row_dimensions[3].height = 22
    ws.freeze_panes = "A4"

    # ── Data rows ──
    selected = votes[:count]
    for row_idx, vote in enumerate(selected, start=1):
        excel_row = row_idx + 3
        fill = alt_fill if row_idx % 2 == 0 else white_fill

        # T/R
        c1 = ws.cell(row=excel_row, column=1, value=row_idx)
        c1.font = cell_font
        c1.fill = fill
        c1.alignment = cell_align_center
        c1.border = border

        # Phone
        phone = vote.get("phoneNumber", vote.get("phone_number", "—"))
        c2 = ws.cell(row=excel_row, column=2, value=phone)
        c2.font = cell_font
        c2.fill = fill
        c2.alignment = cell_align_center
        c2.border = border

        # Date
        vote_date = vote.get("voteDate", vote.get("vote_date", vote.get("date", "—")))
        c3 = ws.cell(row=excel_row, column=3, value=vote_date)
        c3.font = cell_font
        c3.fill = fill
        c3.alignment = cell_align_left
        c3.border = border

        ws.row_dimensions[excel_row].height = 18

    # ── Summary footer ──
    footer_row = len(selected) + 4
    ws.merge_cells(f"A{footer_row}:C{footer_row}")
    footer = ws.cell(row=footer_row, column=1, value=f"Jami: {len(selected)} ta ovoz")
    footer.font = Font(name="Arial", bold=True, size=10, color="1F5C99")
    footer.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[footer_row].height = 20

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ════════════════════════════════════════════
#  COMMANDS
# ════════════════════════════════════════════
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "👋 <b>OpenBudget Ovozlar Boti</b>ga xush kelibsiz!\n\n"
        "Nima qila olaman:\n"
        "📥 <b>/download</b> — Ovozlarni Excel formatda yuklab olish\n"
        "🔍 <b>/search</b> — Telefon raqam bo'yicha ovoz qidirish\n"
        "ℹ️ <b>/help</b> — Yordam\n\n"
        "Boshlash uchun /download yoki /search buyrug'ini yuboring."
    )
    await update.message.reply_text(text, parse_mode="HTML")


async def cmd_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = (
        "📖 <b>Yordam</b>\n\n"
        "<b>/download</b>\n"
        "Tashabbus ID sini kiriting → Nechta ovoz kerakligini tanlang → Excel fayl yuklanadi.\n\n"
        "<b>/search</b>\n"
        "Tashabbus ID sini kiriting → Telefon raqamning oxirgi 6 raqamini kiriting → Natija chiqadi.\n\n"
        "<b>Telefon raqam formati:</b>\n"
        "Saytda raqamlar <code>**-*63-33-35</code> shaklida yashirilgan.\n"
        "Qidiruvda <b>oxirgi 6 raqamni</b> kiriting (masalan: <code>633335</code>)\n\n"
        "❓ Savol bo'lsa, /start yuboring."
    )
    await update.message.reply_text(text, parse_mode="HTML")


# ════════════════════════════════════════════
#  DOWNLOAD FLOW
# ════════════════════════════════════════════
async def cmd_download(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "📋 <b>Tashabbus ID sini kiriting:</b>\n\n"
        "Masalan: <code>0dac589f-08e5-4be8-90c6-9d8435ed497a</code>\n\n"
        "🔗 ID ni openbudget.uz saytidan topishingiz mumkin (URL dagi oxirgi qism).\n\n"
        "Bekor qilish: /cancel",
        parse_mode="HTML",
    )
    return WAITING_INITIATIVE_ID


async def download_got_initiative_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    initiative_id = update.message.text.strip()

    if not is_valid_uuid(initiative_id):
        await update.message.reply_text(
            "❌ Noto'g'ri ID format.\n"
            "UUID ko'rinishida bo'lishi kerak:\n"
            "<code>xxxxxxxx-xxxx-xxxx-xxxx-xxxxxxxxxxxx</code>\n\n"
            "Qaytadan kiriting yoki /cancel",
            parse_mode="HTML",
        )
        return WAITING_INITIATIVE_ID

    context.user_data["initiative_id"] = initiative_id

    msg = await update.message.reply_text("⏳ Ma'lumotlar yuklanmoqda...")

    votes = fetch_votes(initiative_id)
    if votes is None:
        await msg.edit_text(
            "❌ APIdan ma'lumot olishda xatolik yuz berdi.\n"
            "ID to'g'riligini tekshiring va qaytadan urinib ko'ring."
        )
        return ConversationHandler.END

    if len(votes) == 0:
        await msg.edit_text(
            "ℹ️ Bu tashabbus uchun hali ovozlar mavjud emas."
        )
        return ConversationHandler.END

    context.user_data["votes"] = votes
    total = len(votes)

    await msg.delete()

    # Build count selection keyboard
    options = [10, 20, 30, 50, 100]
    keyboard = []
    row = []
    for opt in options:
        label = f"{opt} ta" if opt <= total else f"{opt} ta (jami {total})"
        actual = min(opt, total)
        row.append(InlineKeyboardButton(label, callback_data=f"count_{actual}"))
        if len(row) == 3:
            keyboard.append(row)
            row = []
    if row:
        keyboard.append(row)

    # "Hammasi" button
    keyboard.append([InlineKeyboardButton(f"✅ Hammasi ({total} ta)", callback_data=f"count_{total}")])

    await update.message.reply_text(
        f"✅ Tashabbus topildi!\n"
        f"📊 Jami ovozlar soni: <b>{total} ta</b>\n\n"
        f"Nechta ovoz yuklab olmoqchisiz?",
        reply_markup=InlineKeyboardMarkup(keyboard),
        parse_mode="HTML",
    )
    return WAITING_COUNT


async def download_got_count(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    count = int(query.data.split("_")[1])
    votes = context.user_data.get("votes", [])
    initiative_id = context.user_data.get("initiative_id", "")

    await query.edit_message_text(f"⏳ {count} ta ovoz uchun Excel fayl tayyorlanmoqda...")

    try:
        buf = build_excel(votes, initiative_id, count)
        filename = f"ovozlar_{initiative_id[:8]}_{count}ta_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        await query.message.reply_document(
            document=buf,
            filename=filename,
            caption=(
                f"✅ <b>Excel fayl tayyor!</b>\n"
                f"📊 Tashabbus: <code>{initiative_id}</code>\n"
                f"🗳 Ovozlar soni: <b>{count} ta</b>\n"
                f"📅 Sana: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ),
            parse_mode="HTML",
        )
        await query.delete_message()

    except Exception as e:
        logger.error("Excel build error: %s", e)
        await query.edit_message_text("❌ Excel fayl yaratishda xatolik. Qaytadan urinib ko'ring.")

    context.user_data.clear()
    return ConversationHandler.END


# ════════════════════════════════════════════
#  SEARCH FLOW
# ════════════════════════════════════════════
async def cmd_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "🔍 <b>Ovoz qidirish</b>\n\n"
        "Avval <b>Tashabbus ID</b> sini kiriting:\n"
        "<code>0dac589f-08e5-4be8-90c6-9d8435ed497a</code>\n\n"
        "Bekor qilish: /cancel",
        parse_mode="HTML",
    )
    return WAITING_SEARCH_INITIATIVE


async def search_got_initiative_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    initiative_id = update.message.text.strip()

    if not is_valid_uuid(initiative_id):
        await update.message.reply_text(
            "❌ Noto'g'ri ID format. Qaytadan kiriting yoki /cancel",
            parse_mode="HTML",
        )
        return WAITING_SEARCH_INITIATIVE

    context.user_data["search_initiative_id"] = initiative_id

    msg = await update.message.reply_text("⏳ Ma'lumotlar yuklanmoqda...")
    votes = fetch_votes(initiative_id)

    if votes is None:
        await msg.edit_text("❌ APIdan ma'lumot olishda xatolik. ID ni tekshiring.")
        return ConversationHandler.END

    if len(votes) == 0:
        await msg.edit_text("ℹ️ Bu tashabbus uchun ovozlar mavjud emas.")
        return ConversationHandler.END

    context.user_data["search_votes"] = votes
    await msg.edit_text(
        f"✅ {len(votes)} ta ovoz yuklandi.\n\n"
        f"🔍 Endi <b>telefon raqamning oxirgi 6 raqamini</b> kiriting:\n\n"
        f"Masalan: saytda <code>**-*63-33-35</code> bo'lsa → <code>633335</code> kiriting\n\n"
        f"Bekor qilish: /cancel",
        parse_mode="HTML",
    )
    return WAITING_SEARCH_PHONE


async def search_got_phone(update: Update, context: ContextTypes.DEFAULT_TYPE):
    raw = update.message.text.strip()
    digits = re.sub(r"\D", "", raw)

    if len(digits) < 4 or len(digits) > 11:
        await update.message.reply_text(
            "❌ Raqam noto'g'ri. Oxirgi 4-11 raqamni kiriting.\nMasalan: <code>633335</code>",
            parse_mode="HTML",
        )
        return WAITING_SEARCH_PHONE

    votes = context.user_data.get("search_votes", [])
    initiative_id = context.user_data.get("search_initiative_id", "")
    found = []

    for v in votes:
        phone = v.get("phoneNumber", v.get("phone_number", ""))
        phone_digits = re.sub(r"\D", "", phone)
        if phone_digits.endswith(digits):
            found.append(v)

    if not found:
        await update.message.reply_text(
            f"❌ <b>Topilmadi</b>\n\n"
            f"<code>{digits}</code> raqami bilan tugaydigan ovoz bazada mavjud emas.\n\n"
            f"Tashabbus ID: <code>{initiative_id}</code>\n"
            f"Yana qidirish: /search",
            parse_mode="HTML",
        )
    else:
        lines = [
            f"✅ <b>Topildi! {len(found)} ta ovoz</b>\n\n"
            f"Tashabbus: <code>{initiative_id}</code>\n"
        ]
        for i, v in enumerate(found, 1):
            phone = v.get("phoneNumber", v.get("phone_number", "—"))
            date = v.get("voteDate", v.get("vote_date", v.get("date", "—")))
            lines.append(f"{i}. 📱 <code>{phone}</code>\n   📅 {date}")

        await update.message.reply_text(
            "\n".join(lines) + "\n\nYana qidirish: /search",
            parse_mode="HTML",
        )

    context.user_data.clear()
    return ConversationHandler.END


# ════════════════════════════════════════════
#  CANCEL & FALLBACK
# ════════════════════════════════════════════
async def cmd_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "❌ Bekor qilindi. Bosh menyu: /start"
    )
    return ConversationHandler.END


async def fallback_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Buyruqni tushunmadim. Yordam uchun /help yuboring."
    )


# ════════════════════════════════════════════
#  UTILS
# ════════════════════════════════════════════
UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{4}-[0-9a-f]{12}$",
    re.IGNORECASE,
)

def is_valid_uuid(value: str) -> bool:
    return bool(UUID_RE.match(value.strip()))


# ════════════════════════════════════════════
#  MAIN
# ════════════════════════════════════════════
def main():
    app = Application.builder().token(BOT_TOKEN).build()

    # Download conversation
    download_conv = ConversationHandler(
        entry_points=[CommandHandler("download", cmd_download)],
        states={
            WAITING_INITIATIVE_ID: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, download_got_initiative_id)
            ],
            WAITING_COUNT: [
                CallbackQueryHandler(download_got_count, pattern=r"^count_\d+$")
            ],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
        allow_reentry=True,
    )

    # Search conversation
    search_conv = ConversationHandler(
        entry_points=[CommandHandler("search", cmd_search)],
        states={
            WAITING_SEARCH_INITIATIVE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, search_got_initiative_id)
            ],
            WAITING_SEARCH_PHONE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, search_got_phone)
            ],
        },
        fallbacks=[CommandHandler("cancel", cmd_cancel)],
        allow_reentry=True,
    )

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(download_conv)
    app.add_handler(search_conv)
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, fallback_message))

    logger.info("Bot ishga tushdi...")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
